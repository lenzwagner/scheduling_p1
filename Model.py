import openpyxl
import gurobipy as gu
from gurobipy import *
import seaborn as sns
import pandas as pd
from optimality_plot import plot, parse_log
import matplotlib.pyplot as plt
import time
import os
import csv
import math

# Path
# os.chdir(r'/Users/lorenzwagner/Library/CloudStorage/GoogleDrive-lorenz.wagner99@gmail.com/Meine Ablage/Doktor/Dissertation/Paper 1/Input')
os.chdir(r'G:\Meine Ablage\Doktor\Dissertation\Paper 1\Input')

# Sets
work = pd.read_excel('Arzt.xlsx', sheet_name='Arzt')
df = pd.read_excel('NF.xlsx', sheet_name='NF')
df1 = pd.read_excel('NF.xlsx', sheet_name='Shift')
I = work['Id'].tolist()
W_I = work['Weekend'].tolist()
T = df['Day'].tolist()
K = df1['Shift'].tolist()
S_T = df1['Hours'].tolist()
I_T = work['WT'].tolist()

Min_WD_i = [2, 3, 2, 3]
Max_WD_i = [5, 5, 5, 6]
S_T = {a: c for a, c in zip(K, S_T)}
I_T = {a: d for a, d in zip(I, I_T)}
W_I = {a: e for a, e in zip(I, W_I)}

Min_WD_i = {a: f for a, f in zip(I, Min_WD_i)}
Max_WD_i = {a: g for a, g in zip(I, Max_WD_i)}

I_list1 = pd.DataFrame(I, columns=['I'])
T_list1 = pd.DataFrame(T, columns=['T'])
K_list1 = pd.DataFrame(K, columns=['K'])
DataDF = pd.concat([I_list1, T_list1, K_list1], axis=1)

Demand_Dict = {}
workbook = openpyxl.load_workbook('NF.xlsx')
worksheet = workbook['NF']
for row in worksheet.iter_rows(min_row=2, values_only=True):
    for i in range(1, len(row)):
        cell_value = row[i]
        if isinstance(cell_value, str) and cell_value.startswith('='):
            cell = worksheet.cell(row=row[0], column=i)
            cell_value = cell.value
        if isinstance(cell_value, int):
            Demand_Dict[(int(row[0]), i)] = cell_value
workbook.close()

# Change Dir / remove files
os.chdir(r'G:\Meine Ablage\Doktor\Dissertation\Paper 1\Code\Aktuell')

for file in os.listdir():
    if file.endswith('.log'):
        os.remove(file)

class Problem:
    def __init__(self, dfData, DemandDF, eps, Min_WD, Max_WD):
        self.I = dfData['I'].dropna().astype(int).unique().tolist()
        self.T = dfData['T'].dropna().astype(int).unique().tolist()
        self.K = dfData['K'].dropna().astype(int).unique().tolist()
        self.End = len(self.T)
        self.Week = int(len(self.T) / 7)
        self.Weeks = range(1, self.Week + 1)
        self.demand = DemandDF
        self.model = gu.Model("MasterProblem")
        self.mu = 0.1
        self.epsilon = eps
        self.mue = 0.1
        self.zeta = 0.1
        self.chi = 5
        self.omega = math.floor(1 / 1e-6)
        self.M = len(self.T) + self.omega
        self.xi = 1 - self.epsilon * self.omega
        self.Days_Off = 2
        self.Min_WD = 2
        self.Min_WD_i = Min_WD
        self.Max_WD = 5
        self.Max_WD_i = Max_WD
        self.F_S = [(3, 1), (3, 2), (2, 1)]
        self.objective_values = []
        self.Days = len(self.T)
        self.demand_values = [self.demand[key] for key in self.demand.keys()]

    def buildLinModel(self):
        self.generateVariables()
        self.genGenCons()
        self.genChangesCons()
        self.genRegCons()
        self.Recovery()
        self.linPerformance()
        self.generateObjective()
        self.ModelParams()
        self.updateModel()


    def generateVariables(self):
        self.x = self.model.addVars(self.I, self.T, self.K, vtype=gu.GRB.BINARY, name="x")
        self.kk = self.model.addVars(self.I, self.Weeks, vtype=gu.GRB.BINARY, name="k")
        self.y = self.model.addVars(self.I, self.T, vtype=gu.GRB.BINARY, name="y")
        self.o = self.model.addVars(self.T, self.K, vtype=gu.GRB.CONTINUOUS, name="o")
        self.u = self.model.addVars(self.T, self.K, vtype=gu.GRB.CONTINUOUS, name="u")
        self.sc = self.model.addVars(self.I, self.T, vtype=gu.GRB.BINARY, name="sc")
        self.v = self.model.addVars(self.I, self.T, vtype=gu.GRB.BINARY, name="v")
        self.q = self.model.addVars(self.I, self.T, self.K, vtype=gu.GRB.BINARY, name="q")
        self.rho = self.model.addVars(self.I, self.T, self.K, vtype=gu.GRB.BINARY, name="rho")
        self.z = self.model.addVars(self.I, self.T, self.K, vtype=gu.GRB.BINARY, name="z")
        self.perf = self.model.addVars(self.I, self.T, self.K, vtype=gu.GRB.CONTINUOUS, lb=0, ub=1, name="perf")
        self.p = self.model.addVars(self.I, self.T, vtype=gu.GRB.CONTINUOUS, lb=0, ub=1, name="p")
        self.n = self.model.addVars(self.I, self.T, vtype=gu.GRB.INTEGER, ub=self.Days, lb=0, name="n")
        self.n_h = self.model.addVars(self.I, self.T, vtype=gu.GRB.INTEGER, lb=0, ub=self.Days, name="n_h")
        self.h = self.model.addVars(self.I, self.T, vtype=gu.GRB.BINARY, name="h")
        self.e = self.model.addVars(self.I, self.T, vtype=gu.GRB.BINARY, name="e")
        self.kappa = self.model.addVars(self.I, self.T, vtype=gu.GRB.BINARY, name="kappa")
        self.b = self.model.addVars(self.I, self.T, vtype=gu.GRB.BINARY, name="b")
        self.phi = self.model.addVars(self.I, self.T, vtype=gu.GRB.BINARY, name="phi")
        self.r = self.model.addVars(self.I, self.T, vtype=gu.GRB.BINARY, name="r")
        self.f = self.model.addVars(self.I, self.T, vtype=gu.GRB.BINARY, name="f")
        self.g = self.model.addVars(self.I, self.T, vtype=gu.GRB.BINARY, name="g")
        self.w = self.model.addVars(self.I, self.T, vtype=gu.GRB.BINARY, name="w")
        self.gg = self.model.addVars(self.I, self.T, vtype=gu.GRB.CONTINUOUS, lb=-gu.GRB.INFINITY, ub=gu.GRB.INFINITY,
                                     name="gg")

    def genGenCons(self):
        for i in self.I:
            for t in self.T:
                self.model.addLConstr(gu.quicksum(self.x[i, t, k] for k in self.K) <= 1)
                self.model.addLConstr(gu.quicksum(self.x[i, t, k] for k in self.K) == self.y[i, t])
        for t in self.T:
            for k in self.K:
                self.model.addLConstr(
                    gu.quicksum(self.perf[i, t, k] for i in self.I) + self.u[t, k] >= self.demand[t, k])
        for i in self.I:
            for t in self.T:
                for k in self.K:
                    self.model.addLConstr(self.perf[i, t, k] >= self.p[i, t] - self.M * (1 - self.x[i, t, k]))
                    self.model.addLConstr(self.perf[i, t, k] <= self.p[i, t] + self.M * (1 - self.x[i, t, k]))
                    self.model.addLConstr(self.perf[i, t, k] <= self.x[i, t, k])
        self.model.update()

    def genChangesCons(self):
        for i in self.I:
            for k in self.K:
                for t in self.T:
                    self.model.addLConstr(self.rho[i, t, k] <= 1 - self.q[i, t, k])
                    self.model.addLConstr(self.rho[i, t, k] <= self.x[i, t, k])
                    self.model.addLConstr(self.rho[i, t, k] >= (1 - self.q[i, t, k]) + self.x[i, t, k] - 1)
                    self.model.addLConstr(self.z[i, t, k] <= self.q[i, t, k])
                    self.model.addLConstr(self.z[i, t, k] <= (1 - self.y[i, t]))
                    self.model.addLConstr(self.z[i, t, k] >= self.q[i, t, k] + (1 - self.y[i, t]) - 1)
                for t in range(1, len(self.T)):
                    self.model.addLConstr(self.q[i, t + 1, k] == self.x[i, t, k] + self.z[i, t, k])
            for t in self.T:
                self.model.addLConstr(1 == gu.quicksum(self.x[i, t, k] for k in self.K) + (1 - self.y[i, t]))
                self.model.addLConstr(gu.quicksum(self.rho[i, t, k] for k in self.K) == self.sc[i, t])
        self.model.update()

    def genRegCons(self):
        for i in self.I:
            for t in range(1, len(self.T) - self.Max_WD_i[i] + 1):
                self.model.addLConstr(
                    gu.quicksum(self.y[i, u] for u in range(t, t + 1 + self.Max_WD_i[i])) <= self.Max_WD_i[i])
            for t in range(2, len(self.T) - self.Min_WD_i[i] + 1):
                self.model.addLConstr(
                    gu.quicksum(self.y[i, u] for u in range(t + 1, t + self.Min_WD_i[i] + 1)) >= self.Min_WD_i[i] * (
                                self.y[i, t + 1] - self.y[i, t]))
        for i in self.I:
            for t in range(2, len(self.T) - self.Days_Off + 2):
                for s in range(t + 1, t + self.Days_Off):
                    self.model.addLConstr(1 + self.y[i, t] >= self.y[i, t - 1] + self.y[i, s])
        for i in self.I:
            for k1, k2 in self.F_S:
                for t in range(1, len(self.T)):
                    self.model.addLConstr(self.x[i, t, k1] + self.x[i, t + 1, k2] <= 1)
        self.model.update()

    def Recovery(self):
        for i in self.I:
            for t in range(1 + self.chi, len(self.T) + 1):
                self.model.addLConstr(
                    (1 - self.r[i, t]) <= (1 - self.f[i, t - 1]) + gu.quicksum(self.sc[i, j] for j in range(t - self.chi, t)))
                self.model.addLConstr(self.M * (1 - self.r[i, t]) >= (1 - self.f[i, t - 1]) + gu.quicksum(
                    self.sc[i, j] for j in range(t - self.chi, t)))
            for t in range(1, 1 + self.chi):
                self.model.addLConstr(0 == self.r[i, t])
            for t in self.T:
                for tau in range(1, t + 1):
                    self.model.addLConstr(self.f[i, t] >= self.sc[i, tau])
                self.model.addLConstr(self.f[i, t] <= gu.quicksum(self.sc[i, tau] for tau in range(1, t + 1)))
        self.model.update()

    def linRecovery(self):
        for i in self.I:
            for t in range(2, len(self.T) + 1):
                self.model.addLConstr(self.p[i, t] <= 1)
                self.model.addLConstr(self.p[i, t] <= (1 - self.epsilon * self.n_h[i, t] + self.zeta * self.r[i, t]))
                self.model.addLConstr(self.p[i, t] >= 1 - self.M * self.w[i, t])
                self.model.addLConstr(
                    self.p[i, t] >= (1 - self.epsilon * self.n_h[i, t] + self.zeta * self.r[i, t]) - self.M * (
                                1 - self.w[i, t]))
                self.model.addLConstr(
                    (1 - self.epsilon * self.n_h[i, t] + self.zeta * self.r[i, t]) - 1 <= self.M * self.w[i, t])
                self.model.addLConstr(
                    1 - (1 - self.epsilon * self.n_h[i, t] + self.zeta * self.r[i, t]) <= self.M * (1 - self.w[i, t]))
        self.model.update()


    def linPerformance(self):
        for i in self.I:
            self.model.addLConstr(0 == self.n[i, 1])
            self.model.addLConstr(0 == self.sc[i, 1])
            self.model.addLConstr(1 == self.p[i, 1])
            self.model.addLConstr(0 == self.h[i, 1])
            for t in self.T:
                self.model.addLConstr(
                    gu.quicksum(self.sc[i, j] for j in range(1, t + 1)) <= self.omega + self.M * self.kappa[i, t])
                self.model.addLConstr(self.omega + self.mu <= gu.quicksum(self.sc[i, j] for j in range(1, t + 1)) + (
                            1 - self.kappa[i, t]) * self.M)
            for t in range(2, len(self.T) + 1):
                self.model.addLConstr(self.n[i, t] == self.n_h[i, t] - self.e[i, t] + self.b[i, t])
                self.model.addLConstr(self.n_h[i, t] <= self.n[i, t - 1] + self.sc[i, t])
                self.model.addLConstr(self.n_h[i, t] >= (self.n[i, t - 1] + self.sc[i, t]) - self.M * self.r[i, t])
                self.model.addLConstr(self.n_h[i, t] <= self.M * (1 - self.r[i, t]))
                self.model.addLConstr(self.p[i, t] == 1 - self.epsilon * self.n[i, t] - self.xi * self.kappa[i, t])
                self.model.addLConstr(self.omega * self.h[i, t] <= self.n[i, t])
                self.model.addLConstr(self.n[i, t] <= ((self.omega - 1) + self.h[i, t]))
                self.model.addLConstr(self.e[i, t] <= self.sc[i, t])
                self.model.addLConstr(self.e[i, t] <= self.h[i, t - 1])
                self.model.addLConstr(self.e[i, t] >= self.sc[i, t] + self.h[i, t - 1] - 1)
                self.model.addLConstr(self.b[i, t] <= self.e[i, t])
                self.model.addLConstr(self.b[i, t] <= self.r[i, t])
                self.model.addLConstr(self.b[i, t] >= self.e[i, t] + self.r[i, t] - 1)
        self.model.update()

    def expPerformance(self):
        for i in self.I:
            for t in range(2, len(self.T) + 1):
                self.model.addConstr(self.gg[i, t] == -0.5 * self.n_h[i, t])
                self.model.addGenConstrExp(self.gg[i, t], self.p[i, t], options="FuncNonlinear=1")
        self.model.update()

    def SPerformance(self):
        self.breakpoints = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
        self.slopes = [0] + [-math.exp(-(5 - x)) for x in self.breakpoints[1:]]
        for i in self.I:
            for t in range(2, len(self.T) + 1):
                self.model.addGenConstrPWL(self.gg[i, t], self.n_h[i, t], self.breakpoints, self.slopes)
                self.model.addConstr(self.p[i, t] * (1 + self.gg[i, t]) == 1)
        self.model.update()

    def generateObjective(self):
        self.model.setObjective(gu.quicksum(self.u[t, k] for k in self.K for t in self.T), sense=gu.GRB.MINIMIZE)

    def getObjValues(self):
        obj = self.model.objVal
        return obj

    def updateModel(self):
        self.model.update()

    def ModelParams(self):
        self.model.Params.OutputFlag = 1

    def writeModel(self):
        self.model.write("model.lp")

    def optValues(self):
        values = self.model.getAttr("X", self.x)
        print(values)

    def checkForQuadraticCons(self):
        qconstrs = self.model.getQConstrs()
        print("Check for quadratic constraints:")
        if len(qconstrs) > 0:
            print(True)
        else:
            print(None)

    def calc_u_res(self):
        self.model.Params.OutputFlag = 0
        self.model.optimize()
        u_results = round(sum(self.u[t, k].x for t in self.T for k in self.K), 2)
        return u_results

    def calc_x_res(self):
        self.model.Params.OutputFlag = 0
        self.model.optimize()

        x_values = []
        for i in self.I:
            x_i_values = []
            for t in self.T:
                for k in self.K:
                    x_var = self.x[i, t, k]
                    x_value = self.model.getVarByName(x_var.varName).X
                    x_i_values.append(round(x_value, 3))
            x_values.append(x_i_values)

        sum_xWerte = []
        for i in range(len(x_values[0])):
            sum_value = sum(row[i] for row in x_values)
            sum_xWerte.append(sum_value)

        return sum_xWerte

    def myCallback(self, model, where):
        if where == gu.GRB.Callback.MIPSOL:
            objective_value = model.cbGet(GRB.Callback.MIPSOL_OBJ)
            self.objective_values.append(objective_value)
            cur_obj = model.cbGet(GRB.Callback.MIPSOL_OBJBST)
            cur_bd = model.cbGet(GRB.Callback.MIPSOL_OBJBND)

            if self.model._obj != cur_obj or model._bd != cur_bd:
                self.model._obj = cur_obj
                self.model._bd = cur_bd
                self.model._data.append([time.time() - self.model._start, cur_obj, cur_bd])



    def plotSolution(self, objective_values):
        sns.set(style="darkgrid")

        plt.figure(figsize=(10, 6))

        plt.plot(objective_values, marker='o', linestyle='-')

        plt.xlabel('Time in seconds')
        plt.ylabel('Objective Value')
        plt.title('Objective Function Plot')

        plt.show()

    def saveRuntime(self):
        time = self.model.Runtime
        if time <= 60:
            self.runtime = math.ceil(time)
        else:
            self.runtime = time

        return self.runtime

    def solveModelWithCallback(self):
        try:
            self.model.Params.LogFile = "log1.log"
            self.model._obj = None
            self.model._bd = None
            self.model._data = []
            self.model._start = time.time()
            self.model.optimize(self.myCallback)
            self.saveRuntime()
            print("Final MIP gap value: %f" % self.model.MIPGap)
            self.plotSolution(self.objective_values)
            self.model.write("final.lp")
            if self.model.status == GRB.OPTIMAL:
                print("Optimal solution found")
                for i in self.I:
                    for t in self.T:
                        for k in self.K:
                            if self.x[i, t, k].x > 0.0:
                                print(f"Physician {i}: Shift {k} on day {t}")
            else:
                print("No optimal solution found.")

            print(self.model._data)

            data = self.model._data
            filename = "data.csv"

            with open(filename, 'w', newline='') as file:
                writer = csv.writer(file)
                writer.writerows(data)

        except gu.GurobiError as e:
            print('Error code ' + str(e.errno) + ': ' + str(e))



    def calc_rest(self):
        try:
            self.model.Params.OutputFlag = 0
            self.model.optimize()
            perf_values = []
            for i in self.I:
                perf_i_values = []
                for t in self.T:
                    for k in self.K:
                        perf_var = self.perf[i, t, k]
                        perf_value = self.model.getVarByName(perf_var.varName).X
                        perf_i_values.append(round(perf_value, 3))
                perf_values.append(perf_i_values)
            x_values = []
            for i in self.I:
                x_i_values = []
                for t in self.T:
                    for k in self.K:
                        x_var = self.x[i, t, k]
                        x_value = self.model.getVarByName(x_var.varName).X
                        x_i_values.append(round(x_value, 3))
                x_values.append(x_i_values)
            und_values = self.model.getAttr("X", self.u)
            print(f"x-Values: {x_values}")
            print(f"perf-Values: {perf_values}")
            print(f"u-Values: {und_values}")
        except gu.GurobiError as e:
            print('Error code ' + str(e.errno) + ': ' + str(e))




    def calc_behavior(self, u_results, sum_xWerte):
        try:
            self.sum_xWerte = sum_xWerte
            self.sum_values = sum(self.demand_values)

            self.comparison_result = []
            for i in range(len(self.demand_values)):
                if self.demand_values[i] > self.sum_xWerte[i]:
                    self.comparison_result.append(self.demand_values[i] - self.sum_xWerte[i])
                else:
                    self.comparison_result.append(0)

            self.understaffing = round(sum(self.comparison_result), 3)
            self.perf_loss = round(u_results - self.understaffing, 3)

            print("")
            print(f"Undercoverage: {u_results}")
            print(f"Understaffing: {self.understaffing}")
            print(f"Performance Loss: {self.perf_loss}")
            print("")

            return self.understaffing, self.perf_loss

        except gu.GurobiError as e:
            print('Error code ' + str(e.errno) + ': ' + str(e))

    def calc_naive(self, u_results, sum_xWerte):
        try:
            self.sum_all_doctors = 0
            self.sum_xWerte = sum_xWerte
            self.sum_values = sum(self.demand_values)
            self.cumulative_sum = [0]
            self.doctors_cumulative_multiplied = []
            self.vals = self.demand_values

            self.comp_result = []
            for i in range(len(self.vals)):
                if self.vals[i] < self.sum_xWerte[i]:
                    self.comp_result.append(0)
                else:
                    self.comp_result.append(1)

            self.doctors_cumulative_multiplied = []
            for i in I:
                self.doctor_values = [int(self.sc[i, t].X) for t in self.T]
                self.r_values = []
                for t in self.T:
                    if self.r[i, t].x > 0.5:
                        self.r_values.append(1)
                    else:
                        self.r_values.append(0)
                self.x_i_values = []
                for t in self.T:
                    for k in self.K:
                        if self.x[i, t, k].x > 0.5:
                            self.x_i_values.append(1)
                        else:
                            self.x_i_values.append(0)

                self.cumulative_sum = [0]
                for i in range(1, len(self.doctor_values)):
                    if self.r_values[i] == 1:
                        self.cumulative_sum.append(0)
                    else:
                        self.cumulative_sum.append(self.cumulative_sum[-1] + self.doctor_values[i])

                self.cumulative_sum1 = []
                for element in self.cumulative_sum:
                    for _ in range(len(self.K)):
                        self.cumulative_sum1.append(element)

                self.cumulative_values = [x * self.mue for x in self.cumulative_sum1]
                self.multiplied_values = [self.cumulative_values[j] * self.x_i_values[j] for j in
                                          range(len(self.cumulative_values))]
                self.multiplied_values1 = [self.multiplied_values[j] * self.comp_result[j] for j in
                                           range(len(self.multiplied_values))]
                self.total_sum = sum(self.multiplied_values1)
                self.doctors_cumulative_multiplied.append(self.total_sum)
                self.sum_all_doctors += self.total_sum

            self.understaffing1 = u_results + self.sum_all_doctors
            print(f"Undercoverage: {self.understaffing1}")
            print(f"Understaffing: {u_results}")
            print(f"Performance Loss: {self.sum_all_doctors}")

        except gu.GurobiError as e:
            print('Error code ' + str(e.errno) + ': ' + str(e))

# Build & Solve MP
problem = Problem(DataDF, Demand_Dict, 0.2, Min_WD_i, Max_WD_i)
problem.buildLinModel()
problem.updateModel()
problem.checkForQuadraticCons()
problem.solveModelWithCallback()
model_time = problem.model.Runtime

problem.calc_behavior(problem.calc_u_res(), problem.calc_x_res())
problem.calc_naive(problem.calc_u_res(), problem.calc_x_res())


# Optimality Plots
file_path = "./log1.log"
plot(parse_log(file_path), problem.runtime)